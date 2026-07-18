from __future__ import annotations

import json
import re
import unicodedata
from datetime import date, datetime
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill("solid", fgColor="0F4C81")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="0F4C81", bold=True, size=16)
SECTION_FONT = Font(color="17212B", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="DCE4EA"),
    right=Side(style="thin", color="DCE4EA"),
    top=Side(style="thin", color="DCE4EA"),
    bottom=Side(style="thin", color="DCE4EA"),
)
CURRENCY_FORMAT = 'R$ #,##0.00'
DECIMAL_FORMAT = '#,##0.000'
DATE_FORMAT = 'dd/mm/yyyy'
DATETIME_FORMAT = 'dd/mm/yyyy hh:mm'


def _slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = "".join(character for character in normalized if not unicodedata.combining(character))
    compact = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_value).strip("-").lower()
    return compact[:60] or "familia"


def _parse_temporal(value: Any) -> Any:
    if not isinstance(value, str) or not value:
        return value
    try:
        if "T" in value or " " in value:
            return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
        return date.fromisoformat(value)
    except ValueError:
        return value


def _safe(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False, separators=(",", ":"))

    parsed = _parse_temporal(value)
    if isinstance(parsed, str) and parsed.startswith(("=", "+", "-", "@")):
        return "'" + parsed
    return parsed


def _add_table_sheet(
    workbook: Workbook,
    *,
    title: str,
    table_name: str,
    headers: list[str],
    rows: Iterable[Iterable[Any]],
    currency_columns: set[int] | None = None,
    decimal_columns: set[int] | None = None,
    date_columns: set[int] | None = None,
    datetime_columns: set[int] | None = None,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)

    row_count = 1
    for row in rows:
        sheet.append([_safe(value) for value in row])
        row_count += 1

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for index in currency_columns or set():
        for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
            for item in cell:
                item.number_format = CURRENCY_FORMAT

    for index in decimal_columns or set():
        for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
            for item in cell:
                item.number_format = DECIMAL_FORMAT

    for index in date_columns or set():
        for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
            for item in cell:
                item.number_format = DATE_FORMAT

    for index in datetime_columns or set():
        for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
            for item in cell:
                item.number_format = DATETIME_FORMAT

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    if row_count >= 2:
        table = Table(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    for column_cells in sheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 11), 36)

    sheet.row_dimensions[1].height = 24


def criar_excel_exportacao(payload: dict[str, Any]) -> tuple[bytes, str]:
    family = payload.get("familia") if isinstance(payload.get("familia"), dict) else {}
    configuration = payload.get("configuracoes") if isinstance(payload.get("configuracoes"), dict) else {}
    family_name = str(family.get("nome") or "Família")
    generated_at = _parse_temporal(payload.get("gerado_em"))
    timestamp = generated_at if isinstance(generated_at, datetime) else datetime.now()

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary["A1"] = "Gestão de Compras — Exportação da família"
    summary["A1"].font = TITLE_FONT
    summary.merge_cells("A1:D1")

    summary_rows = [
        ("Família", family_name),
        ("Plano", str(family.get("plano") or "free").title()),
        ("Status", str(family.get("status") or "ativa").title()),
        ("Gerado em", timestamp),
        ("Moeda", configuration.get("moeda") or "BRL"),
        ("Localidade", configuration.get("localidade") or "pt-BR"),
        ("Membros", len(payload.get("membros") or [])),
        ("Compras", len(payload.get("compras") or [])),
        ("Itens", len(payload.get("itens_compra") or [])),
        ("Produtos", len(payload.get("produtos") or [])),
        ("Registros de preço", len(payload.get("historico_precos") or [])),
        ("Supermercados", len(payload.get("supermercados") or [])),
        ("Categorias", len(payload.get("categorias") or [])),
    ]

    for row_index, (label, value) in enumerate(summary_rows, start=3):
        summary.cell(row=row_index, column=1, value=label).font = SECTION_FONT
        summary.cell(row=row_index, column=2, value=_safe(value))
        summary.cell(row=row_index, column=1).border = THIN_BORDER
        summary.cell(row=row_index, column=2).border = THIN_BORDER

    summary["B6"].number_format = DATETIME_FORMAT
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 34
    summary.freeze_panes = "A3"

    members = payload.get("membros") or []
    _add_table_sheet(
        workbook,
        title="Membros",
        table_name="MembrosTable",
        headers=["Nome", "E-mail", "Papel", "Status", "Criado em"],
        rows=[
            [item.get("nome"), item.get("email"), item.get("papel"), item.get("status"), item.get("criado_em")]
            for item in members
        ],
        datetime_columns={5},
    )

    purchases = payload.get("compras") or []
    _add_table_sheet(
        workbook,
        title="Compras",
        table_name="ComprasTable",
        headers=[
            "Data", "Supermercado", "CNPJ/Chave NFC-e", "Valor total", "Valor pago",
            "Forma de pagamento", "Origem", "Status", "Registrado por", "Criado em", "ID",
        ],
        rows=[
            [
                item.get("data_compra"), item.get("supermercado_nome"), item.get("chave_nfce"),
                item.get("valor_total"), item.get("valor_pago"), item.get("forma_pagamento"),
                item.get("origem"), item.get("status"), item.get("criado_por_email"),
                item.get("criado_em"), item.get("id"),
            ]
            for item in purchases
        ],
        currency_columns={4, 5},
        date_columns={1},
        datetime_columns={10},
    )

    items = payload.get("itens_compra") or []
    _add_table_sheet(
        workbook,
        title="Itens",
        table_name="ItensTable",
        headers=[
            "Data", "Supermercado", "Produto", "Descrição original", "Categoria", "Quantidade",
            "Unidade", "Valor unitário", "Valor total", "Compra ID", "Produto ID",
        ],
        rows=[
            [
                item.get("data_compra"), item.get("supermercado_nome"), item.get("produto_nome"),
                item.get("descricao_original"), item.get("categoria_nome"), item.get("quantidade"),
                item.get("unidade"), item.get("valor_unitario"), item.get("valor_total"),
                item.get("compra_id"), item.get("produto_id"),
            ]
            for item in items
        ],
        currency_columns={8, 9},
        decimal_columns={6},
        date_columns={1},
    )

    products = payload.get("produtos") or []
    _add_table_sheet(
        workbook,
        title="Produtos",
        table_name="ProdutosTable",
        headers=["Produto", "Marca", "Unidade", "Categoria", "Revisar", "Ativo", "Criado em", "Atualizado em", "ID"],
        rows=[
            [
                item.get("nome"), item.get("marca"), item.get("unidade_padrao"), item.get("categoria_nome"),
                "Sim" if item.get("revisar") else "Não", "Sim" if item.get("ativo") else "Não",
                item.get("criado_em"), item.get("atualizado_em"), item.get("id"),
            ]
            for item in products
        ],
        datetime_columns={7, 8},
    )

    history = payload.get("historico_precos") or []
    _add_table_sheet(
        workbook,
        title="Histórico preços",
        table_name="HistoricoPrecosTable",
        headers=[
            "Data", "Produto", "Categoria", "Supermercado", "Quantidade", "Unidade",
            "Valor unitário", "Compra ID", "Produto ID",
        ],
        rows=[
            [
                item.get("data_compra"), item.get("produto_nome"), item.get("categoria_nome"),
                item.get("supermercado_nome"), item.get("quantidade"), item.get("unidade"),
                item.get("valor_unitario"), item.get("compra_id"), item.get("produto_id"),
            ]
            for item in history
        ],
        currency_columns={7},
        decimal_columns={5},
        date_columns={1},
    )

    supermarkets = payload.get("supermercados") or []
    _add_table_sheet(
        workbook,
        title="Supermercados",
        table_name="SupermercadosTable",
        headers=["Nome", "CNPJ", "Ativo", "Criado em", "Atualizado em", "ID"],
        rows=[
            [
                item.get("nome"), item.get("cnpj"), "Sim" if item.get("ativo") else "Não",
                item.get("criado_em"), item.get("atualizado_em"), item.get("id"),
            ]
            for item in supermarkets
        ],
        datetime_columns={4, 5},
    )

    categories = payload.get("categorias") or []
    _add_table_sheet(
        workbook,
        title="Categorias",
        table_name="CategoriasTable",
        headers=["Nome", "Categoria do sistema", "Ativa", "Criado em", "Atualizado em", "ID"],
        rows=[
            [
                item.get("nome"), "Sim" if item.get("sistema") else "Não",
                "Sim" if item.get("ativo") else "Não", item.get("criado_em"),
                item.get("atualizado_em"), item.get("id"),
            ]
            for item in categories
        ],
        datetime_columns={4, 5},
    )

    workbook.properties.title = f"Gestão de Compras — {family_name}"
    workbook.properties.subject = "Exportação completa dos dados da família"
    workbook.properties.creator = "Gestão de Compras Web"

    output = BytesIO()
    workbook.save(output)
    filename = f"gestao-compras-{_slug(family_name)}-{timestamp:%Y-%m-%d}.xlsx"
    return output.getvalue(), filename


def criar_backup_json(payload: dict[str, Any]) -> tuple[bytes, str]:
    family = payload.get("familia") if isinstance(payload.get("familia"), dict) else {}
    family_name = str(family.get("nome") or "Família")
    generated_at = _parse_temporal(payload.get("gerado_em"))
    timestamp = generated_at if isinstance(generated_at, datetime) else datetime.now()
    content = json.dumps(payload, ensure_ascii=False, indent=2, default=str).encode("utf-8")
    filename = f"backup-gestao-compras-{_slug(family_name)}-{timestamp:%Y-%m-%d}.json"
    return content, filename
