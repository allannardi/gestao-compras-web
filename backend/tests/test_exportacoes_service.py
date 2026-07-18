import json
from io import BytesIO

from openpyxl import load_workbook

from app.services.exportacoes import criar_backup_json, criar_excel_exportacao


def _payload() -> dict:
    return {
        "formato": "gestao-compras-backup",
        "versao_schema": 12,
        "gerado_em": "2026-07-18T12:00:00+00:00",
        "familia": {
            "id": "family-1",
            "nome": "Família Nardi",
            "plano": "free",
            "status": "ativa",
        },
        "configuracoes": {"moeda": "BRL", "localidade": "pt-BR"},
        "membros": [
            {
                "nome": "Allan",
                "email": "allan@example.com",
                "papel": "administrador",
                "status": "ativo",
                "criado_em": "2026-07-01T10:00:00+00:00",
            }
        ],
        "categorias": [
            {
                "id": "category-1",
                "nome": "Alimentos básicos",
                "sistema": True,
                "ativo": True,
                "criado_em": "2026-07-01T10:00:00+00:00",
                "atualizado_em": "2026-07-01T10:00:00+00:00",
            }
        ],
        "supermercados": [
            {
                "id": "market-1",
                "nome": "Mercado Central",
                "cnpj": "12345678000199",
                "ativo": True,
                "criado_em": "2026-07-01T10:00:00+00:00",
                "atualizado_em": "2026-07-01T10:00:00+00:00",
            }
        ],
        "produtos": [
            {
                "id": "product-1",
                "nome": "Arroz",
                "marca": "Marca A",
                "unidade_padrao": "un",
                "categoria_nome": "Alimentos básicos",
                "revisar": False,
                "ativo": True,
                "criado_em": "2026-07-01T10:00:00+00:00",
                "atualizado_em": "2026-07-01T10:00:00+00:00",
            }
        ],
        "compras": [
            {
                "id": "purchase-1",
                "data_compra": "2026-07-17",
                "supermercado_nome": "Mercado Central",
                "chave_nfce": "1" * 44,
                "valor_total": 25.5,
                "valor_pago": 25.5,
                "forma_pagamento": "Cartão",
                "origem": "nfce_qrcode",
                "status": "confirmada",
                "criado_por_email": "allan@example.com",
                "criado_em": "2026-07-17T20:00:00+00:00",
            }
        ],
        "itens_compra": [
            {
                "data_compra": "2026-07-17",
                "supermercado_nome": "Mercado Central",
                "produto_nome": "Arroz",
                "descricao_original": "ARROZ 5KG",
                "categoria_nome": "Alimentos básicos",
                "quantidade": 1,
                "unidade": "un",
                "valor_unitario": 25.5,
                "valor_total": 25.5,
                "compra_id": "purchase-1",
                "produto_id": "product-1",
            }
        ],
        "historico_precos": [
            {
                "data_compra": "2026-07-17",
                "produto_nome": "Arroz",
                "categoria_nome": "Alimentos básicos",
                "supermercado_nome": "Mercado Central",
                "quantidade": 1,
                "unidade": "un",
                "valor_unitario": 25.5,
                "compra_id": "purchase-1",
                "produto_id": "product-1",
            }
        ],
    }


def test_cria_excel_com_abas_e_formatos() -> None:
    content, filename = criar_excel_exportacao(_payload())

    workbook = load_workbook(BytesIO(content), data_only=False)
    assert filename == "gestao-compras-familia-nardi-2026-07-18.xlsx"
    assert workbook.sheetnames == [
        "Resumo",
        "Membros",
        "Compras",
        "Itens",
        "Produtos",
        "Histórico preços",
        "Supermercados",
        "Categorias",
    ]
    assert workbook["Compras"]["A2"].value.date().isoformat() == "2026-07-17"
    assert workbook["Compras"]["D2"].number_format == "R$ #,##0.00"
    assert workbook["Itens"]["C2"].value == "Arroz"
    assert workbook["Resumo"]["B3"].value == "Família Nardi"


def test_excel_neutraliza_formula_em_texto() -> None:
    payload = _payload()
    payload["produtos"][0]["nome"] = "=HYPERLINK(\"https://exemplo.com\")"
    content, _ = criar_excel_exportacao(payload)
    workbook = load_workbook(BytesIO(content), data_only=False)
    assert workbook["Produtos"]["A2"].value.startswith("'=")


def test_cria_backup_json_legivel() -> None:
    content, filename = criar_backup_json(_payload())
    parsed = json.loads(content.decode("utf-8"))

    assert filename == "backup-gestao-compras-familia-nardi-2026-07-18.json"
    assert parsed["versao_schema"] == 12
    assert parsed["compras"][0]["valor_total"] == 25.5
